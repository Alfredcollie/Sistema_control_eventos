# -*- coding: utf-8 -*-

import psycopg2
import tkinter as tk
import urllib.request
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
from datetime import datetime
import ctypes
import calendar
import sys
import subprocess
import os
import re
import json
import threading
import importlib
import unicodedata

# 🚀 IMPORTAMOS NUESTRAS NUEVAS HERRAMIENTAS CORPORATIVAS
from conexion import conectar_db, registrar_auditoria, liberar_conexion
from buffer_memoria import cache_sistema

try:
    from app_paths import CONFIG_FILE
    RUTA_CONFIG = str(CONFIG_FILE)
except Exception:
    RUTA_CONFIG = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config_local.json")

ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("blue")

COLOR_PRIMARIO = "#eb337a"

# =========================================================
# MULTIPLATAFORMA: Funciones Universales
# =========================================================
if sys.platform == "win32":
    try:
        hwnd_cmd = ctypes.windll.kernel32.GetConsoleWindow()
        if hwnd_cmd:
            ctypes.windll.user32.ShowWindow(hwnd_cmd, 6)
    except Exception:
        pass

def maximizar_ventana(ventana):
    try:
        if sys.platform == "win32":
            ventana.state("zoomed")
        elif sys.platform == "darwin":
            ventana.attributes("-zoomed", True)
        else:
            ventana.state("zoomed")
    except Exception:
        try:
            w = ventana.winfo_screenwidth()
            h = ventana.winfo_screenheight()
            ventana.geometry(f"{w}x{h}+0+0")
        except Exception:
            pass

def abrir_documento(ruta):
    """Abre el archivo generado con el visor predeterminado según el Sistema Operativo"""
    try:
        if sys.platform == "win32":
            os.startfile(ruta)
        elif sys.platform == "darwin":
            subprocess.call(["open", ruta])
        else:
            subprocess.call(["xdg-open", ruta])
    except Exception as e:
        print(f"No se pudo abrir el documento automáticamente: {e}")

# =========================================================
# HERRAMIENTAS DE TEXTO ENRIQUECIDO - VERSIÓN WYSIWYG MAC FIX
# =========================================================
_PATRON_ETIQUETAS = re.compile(r'(\[B\]|\[/B\]|\[M\]|\[/M\])', re.IGNORECASE)

def parsear_segmentos_formato(texto):
    resultado, negrita, color_p = [], False, False
    for parte in _PATRON_ETIQUETAS.split(str(texto)):
        p_up = parte.upper()
        if p_up == "[B]":
            negrita = True
        elif p_up == "[/B]":
            negrita = False
        elif p_up == "[M]":
            color_p = True
        elif p_up == "[/M]":
            color_p = False
        elif parte:
            resultado.append((parte, negrita, color_p))
    return resultado

def texto_plano_sin_marcado(texto):
    return _PATRON_ETIQUETAS.sub("", str(texto))

# Memoria de normalización (perf): evita recalcular el texto normalizado de cada
# proveedor en cada cambio de categoría dentro de la misma sesión.
_NORM_CACHE_PROVEEDORES = {}

def extraer_texto_con_formato(txt_widget):
    inner = txt_widget._textbox if hasattr(txt_widget, "_textbox") else txt_widget
    dump = inner.dump("1.0", "end-1c")
    partes = []
    for key, value, index in dump:
        if key == "tagon":
            if value == "bold": partes.append("[B]")
            elif value == "color": partes.append("[M]")
        elif key == "tagoff":
            if value == "bold": partes.append("[/B]")
            elif value == "color": partes.append("[/M]")
        elif key == "text":
            partes.append(value)
    return "".join(partes)

def crear_barra_formato(parent, text_widget):
    f_barra = ctk.CTkFrame(parent, fg_color="transparent")
    
    inner_text = text_widget._textbox if hasattr(text_widget, "_textbox") else text_widget
    inner_text._memoria_blindada = None
    inner_text._memoria_bloqueada = False

    def rastreador_mac():
        try:
            if not inner_text.winfo_exists(): return
            if inner_text.tag_ranges(tk.SEL):
                inner_text._memoria_blindada = (inner_text.index(tk.SEL_FIRST), inner_text.index(tk.SEL_LAST))
            elif not inner_text._memoria_bloqueada:
                inner_text._memoria_blindada = None
        except Exception:
            pass
        inner_text.after(50, rastreador_mac)

    rastreador_mac()

    def activar_candado(e): inner_text._memoria_bloqueada = True
    def quitar_candado(e): inner_text._memoria_bloqueada = False

    f_barra.bind("<Enter>", activar_candado, add="+")
    f_barra.bind("<Leave>", quitar_candado, add="+")

    def alternar_formato(tag_name):
        s, e = None, None
        try:
            if inner_text.tag_ranges(tk.SEL):
                s = inner_text.index(tk.SEL_FIRST)
                e = inner_text.index(tk.SEL_LAST)
        except Exception:
            pass

        if not s and inner_text._memoria_blindada:
            s, e = inner_text._memoria_blindada

        if s and e:
            try:
                current_tags = inner_text.tag_names(s)
                if tag_name in current_tags:
                    inner_text.tag_remove(tag_name, s, e)
                else:
                    inner_text.tag_add(tag_name, s, e)
                
                inner_text.tag_raise(tag_name)
                
                inner_text.tag_add(tk.SEL, s, e)
                inner_text._memoria_blindada = None
            except Exception:
                pass
                
        inner_text.focus_set()
        return "break"

    btn_b = ctk.CTkLabel(f_barra, text=" B ", width=30, height=25, font=("Helvetica", 12, "bold"), fg_color="#e0e0e0", text_color="black", corner_radius=5, cursor="hand2")
    btn_b.pack(side="left", padx=2)
    btn_b.bind("<Button-1>", lambda e: alternar_formato("bold"))
    btn_b.bind("<Enter>", lambda e: [activar_candado(e), btn_b.configure(fg_color="#c8c8c8")])
    btn_b.bind("<Leave>", lambda e: [quitar_candado(e), btn_b.configure(fg_color="#e0e0e0")])
    
    btn_c = ctk.CTkLabel(f_barra, text=" Color ", width=45, height=25, font=("Helvetica", 12, "bold"), fg_color="#e0e0e0", text_color=COLOR_PRIMARIO, corner_radius=5, cursor="hand2")
    btn_c.pack(side="left", padx=2)
    btn_c.bind("<Button-1>", lambda e: alternar_formato("color"))
    btn_c.bind("<Enter>", lambda e: [activar_candado(e), btn_c.configure(fg_color="#c8c8c8")])
    btn_c.bind("<Leave>", lambda e: [quitar_candado(e), btn_c.configure(fg_color="#e0e0e0")])

    inner_text.bind("<Command-b>", lambda e: alternar_formato("bold"))
    inner_text.bind("<Command-m>", lambda e: alternar_formato("color"))
    inner_text.bind("<Control-b>", lambda e: alternar_formato("bold"))
    inner_text.bind("<Control-m>", lambda e: alternar_formato("color"))
    
    configurar_tags_formato(text_widget, tam=11)
    
    return f_barra

def configurar_tags_formato(txt_widget, tam=10):
    inner = txt_widget._textbox if hasattr(txt_widget, "_textbox") else txt_widget
    inner.tag_configure("bold", font=("Helvetica", tam, "bold"))
    inner.tag_configure("color", foreground=COLOR_PRIMARIO)
    inner.tag_raise("bold")
    inner.tag_raise("color")

def insertar_texto_formateado(txt_widget, texto):
    inner = txt_widget._textbox if hasattr(txt_widget, "_textbox") else txt_widget
    inner.delete("1.0", tk.END)
    segmentos = parsear_segmentos_formato(texto)
    for frag, neg, col in segmentos:
        tags = []
        if neg: tags.append("bold")
        if col: tags.append("color")
        
        if tags:
            inner.insert(tk.END, frag, tuple(tags))
        else:
            inner.insert(tk.END, frag)
            
    inner.tag_raise("bold")
    inner.tag_raise("color")

# =========================================================
# 🚀 FUNCIONES GENERADORAS DE CÓDIGOS CORRELATIVOS
# =========================================================
def generar_nuevo_codigo_cotizacion(conn):
    prefijo_fecha = datetime.now().strftime("%y%m%d")
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT codigo_cotizacion FROM cotizaciones WHERE codigo_cotizacion LIKE %s ORDER BY id DESC LIMIT 1", (f"{prefijo_fecha}-%",))
        res = cursor.fetchone()
        if res and res[0]:
            partes = str(res[0]).split('-')
            if len(partes) >= 2:
                try:
                    secuencial = int(partes[1]) + 1
                except ValueError:
                    secuencial = 1
            else:
                secuencial = 1
        else:
            secuencial = 1
        return f"{prefijo_fecha}-{secuencial:02d}-01"
    except Exception as e:
        print("Error generando código de cotización:", e)
        return f"{prefijo_fecha}-01-01"

def generar_nueva_version_evento_existente(conn, codigo_actual):
    partes = str(codigo_actual).split('-')
    if len(partes) >= 2:
        base = f"{partes[0]}-{partes[1]}"
    else:
        base = str(codigo_actual)
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT codigo_cotizacion FROM cotizaciones WHERE codigo_cotizacion LIKE %s ORDER BY id DESC LIMIT 1", (f"{base}-%",))
        res = cursor.fetchone()
        if res and res[0]:
            partes_ult = str(res[0]).split('-')
            if len(partes_ult) == 3:
                try:
                    version = int(partes_ult[2]) + 1
                except ValueError:
                    version = 2
            else:
                version = 2
        else:
            version = 2
        return f"{base}-{version:02d}"
    except Exception as e:
        print("Error generando versión de cotización:", e)
        return f"{base}-02"

# =========================================================
# CLASE: CALENDARIO NATIVO
# =========================================================
class CalendarioNativo(ctk.CTkToplevel):
    def __init__(self, parent, target_entry):
        super().__init__(parent)
        self.target_entry = target_entry
        self.title("Seleccionar Fecha")
        self.geometry("310x320")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() // 2) - (310 // 2)
        y = parent.winfo_rooty() + (parent.winfo_height() // 2) - (320 // 2)
        self.geometry(f"+{x}+{y}")

        ahora = datetime.now()
        self.current_year = ahora.year
        self.current_month = ahora.month
        self.meses_nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

        self.header_frame = ctk.CTkFrame(self, fg_color="#1f538d", corner_radius=0)
        self.header_frame.pack(fill="x")

        ctk.CTkButton(self.header_frame, text="<", width=25, fg_color="transparent", text_color="white", hover_color="#163b65", font=("Arial", 14, "bold"), command=self.prev_month).pack(side="left", padx=5, pady=10)

        self.cmb_mes = ctk.CTkComboBox(self.header_frame, values=self.meses_nombres, width=100, command=self.cambiar_mes_combo)
        self.cmb_mes.pack(side="left", padx=2, pady=10)

        anio_actual = datetime.now().year
        anios = [str(y) for y in range(anio_actual - 80, anio_actual + 20)]
        self.cmb_anio = ctk.CTkComboBox(self.header_frame, values=anios, width=75, command=self.cambiar_anio_combo)
        self.cmb_anio.pack(side="left", padx=2, pady=10)

        ctk.CTkButton(self.header_frame, text=">", width=25, fg_color="transparent", text_color="white", hover_color="#163b65", font=("Arial", 14, "bold"), command=self.next_month).pack(side="right", padx=5, pady=10)

        self.days_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.days_frame.pack(fill="both", expand=True, padx=10, pady=10)

        dias_semana = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
        for i, day in enumerate(dias_semana):
            ctk.CTkLabel(self.days_frame, text=day, font=("Arial", 11, "bold"), text_color="#1f538d").grid(row=0, column=i, padx=5, pady=5)

        self.update_calendar()

    def cambiar_mes_combo(self, choice):
        self.current_month = self.meses_nombres.index(choice) + 1
        self.update_calendar()

    def cambiar_anio_combo(self, choice):
        try:
            self.current_year = int(choice)
            self.update_calendar()
        except ValueError:
            pass

    def update_calendar(self):
        self.cmb_mes.set(self.meses_nombres[self.current_month - 1])
        self.cmb_anio.set(str(self.current_year))

        for widget in self.days_frame.winfo_children():
            if int(widget.grid_info()["row"]) > 0: widget.destroy()

        cal = calendar.monthcalendar(self.current_year, self.current_month)
        hoy = datetime.now()

        for row_idx, week in enumerate(cal, start=1):
            for col_idx, day in enumerate(week):
                if day != 0:
                    btn_color = "#d4edda" if day == hoy.day and self.current_month == hoy.month and self.current_year == hoy.year else "transparent"
                    txt_color = "#155724" if btn_color == "#d4edda" else "black"
                    btn = ctk.CTkButton(self.days_frame, text=str(day), width=30, height=30, fg_color=btn_color, text_color=txt_color, hover_color="#e0e0e0", font=("Arial", 11))
                    btn.configure(command=lambda d=day: self.select_date(d))
                    btn.grid(row=row_idx, column=col_idx, padx=3, pady=2)

    def prev_month(self):
        self.current_month -= 1
        if self.current_month < 1: self.current_month = 12; self.current_year -= 1
        self.update_calendar()

    def next_month(self):
        self.current_month += 1
        if self.current_month > 12: self.current_month = 1; self.current_year += 1
        self.update_calendar()

    def select_date(self, day):
        fecha_seleccionada = f"{day:02d}/{self.current_month:02d}/{self.current_year}"
        self.target_entry.delete(0, tk.END)
        self.target_entry.insert(0, fecha_seleccionada)
        self.destroy()


# =======================================================
# CLASE PRINCIPAL - ETAPA 3 (MATRIZ DE PROVEEDORES)
# =======================================================
_SCHEMA_F3_OK = False

class VentanaEtapaProveedores:
    def __init__(self, parent_ventana, codigo_cot, empresa, evento, callback_on_close=None):
        self.parent_ventana = parent_ventana
        self.root = parent_ventana.root if hasattr(parent_ventana, 'root') else parent_ventana
        self.conn = conectar_db()
        self.codigo_cot = str(codigo_cot).strip()
        self.empresa = empresa
        self.evento = evento
        self.callback_on_close = callback_on_close
        self.usuario_activo = getattr(self.parent_ventana, 'usuario_activo', 'Desconocido')
        ctk.set_appearance_mode("Light")

        if not self.conn:
            messagebox.showwarning("Sin conexión", "No hay conexión con la base de datos.\nLa Etapa 3 no puede abrirse en Modo Lectura.", parent=self.root)
            self.v_prov = None
            return

        self.v_prov = ctk.CTkToplevel(self.root)
        self.v_prov.title(f"Etapa 3: Matriz de Costos - Cotización: {self.codigo_cot}")
        self.v_prov.geometry("1200x780")
        self.v_prov.grab_set()
        self.v_prov.after(100, lambda: maximizar_ventana(self.v_prov))
        self.v_prov.protocol("WM_DELETE_WINDOW", self._cerrar_ventana)

        # Scroll de rueda global sobre la matriz y las notas (idempotente: solo
        # se instala una vez por sesión). Vive en scroll_utils.py.
        try:
            from scroll_utils import instalar_scroll_global
            instalar_scroll_global(self.v_prov)
        except Exception:
            pass

        self.fila_matriz_seleccionada = None
        self.lista_widgets_filas = []
        self.matriz_expandida = False

        global _SCHEMA_F3_OK
        if not _SCHEMA_F3_OK:
            def tarea_init():
                global _SCHEMA_F3_OK
                c_conn = conectar_db(silencioso=True)
                if c_conn:
                    try:
                        c_alt = c_conn.cursor()
                        alters = [
                            "ALTER TABLE cotizaciones ADD COLUMN IF NOT EXISTS tipo_cambio NUMERIC DEFAULT 3.75",
                            "ALTER TABLE cotizaciones ADD COLUMN IF NOT EXISTS forma_pago TEXT DEFAULT '50% adelantado, 50% a 30 días de la primera factura.'",
                            "ALTER TABLE cotizaciones ADD COLUMN IF NOT EXISTS sin_fee BOOLEAN DEFAULT FALSE",
                            "ALTER TABLE cotizacion_proveedores ADD COLUMN IF NOT EXISTS cantidad INTEGER DEFAULT 1",
                            "ALTER TABLE cotizacion_proveedores ADD COLUMN IF NOT EXISTS notas_internas TEXT DEFAULT ''",
                        ]
                        for sql in alters:
                            try:
                                c_alt.execute(sql)
                                c_conn.commit()
                            except: c_conn.rollback()
                        _SCHEMA_F3_OK = True
                    except Exception: 
                        pass
                    finally: 
                        liberar_conexion(c_conn)
            threading.Thread(target=tarea_init, daemon=True).start()

        self.f_info = ctk.CTkFrame(self.v_prov, corner_radius=10, fg_color="#1f538d")
        self.f_info.pack(fill="x", padx=15, pady=(15, 5))
        ctk.CTkLabel(self.f_info, text=f"Cliente: {self.empresa}   |   Evento: {self.evento}   |   N° Cotización: {self.codigo_cot}", font=("Arial", 12, "bold"), text_color="white").pack(anchor="w", padx=15, pady=8)

        self.f_inputs = ctk.CTkFrame(self.v_prov, corner_radius=10)
        self.f_inputs.pack(fill="x", padx=15, pady=5)
        self.f_inputs.grid_columnconfigure(4, weight=1)
        ctk.CTkLabel(self.f_inputs, text="Registrar Costo y Margen Comercial por Categoría", font=("Arial", 14, "bold"), text_color="#1f538d").grid(row=0, column=0, columnspan=8, sticky="w", padx=15, pady=(10, 5))

        ctk.CTkLabel(self.f_inputs, text="Seleccione Categoría:", font=("Arial", 12, "bold")).grid(row=1, column=0, sticky="w", pady=5, padx=(15, 5))
        self.cats_assigned = []
        try:
            c = self.conn.cursor()
            c.execute("SELECT categoria_suministro FROM cotizacion_detalles WHERE codigo_cotizacion = %s", (self.codigo_cot,))
            for row in c.fetchall():
                txt_cat = str(row[0]).replace("('", "").replace("',)", "").replace("',", "").strip("() '\", ")
                if txt_cat and txt_cat not in self.cats_assigned:
                    self.cats_assigned.append(txt_cat)
        except Exception:
            pass
        if not self.cats_assigned:
            self.cats_assigned = ["No hay categorias disponibles"]
            
        self.cmb_cat_e = ctk.CTkComboBox(self.f_inputs, values=self.cats_assigned, state="readonly", width=180, command=self.filtrar_proveedores_por_categoria)
        self.cmb_cat_e.grid(row=1, column=1, sticky="w", pady=5, padx=5)
        self.cmb_cat_e.set(self.cats_assigned[0])

        ctk.CTkLabel(self.f_inputs, text="Proveedor asignado:", font=("Arial", 12, "bold")).grid(row=1, column=2, sticky="w", pady=5, padx=15)
        f_prov_accion = ctk.CTkFrame(self.f_inputs, fg_color="transparent")
        f_prov_accion.grid(row=1, column=3, sticky="w", pady=5, padx=5)
        
        self.cmb_p_list = ctk.CTkComboBox(f_prov_accion, values=["--- Seleccione Proveedor ---"], state="readonly", width=250)
        self.cmb_p_list.pack(side="left", padx=(0, 5))
        self.cmb_p_list.set("--- Seleccione Proveedor ---")

        def abrir_sistema_proveedores():
            try:
                import proveedores
                v_ext = ctk.CTkToplevel(self.v_prov)
                v_ext.after(100, lambda: maximizar_ventana(v_ext))
                v_ext.transient(self.v_prov)
                v_ext.grab_set()
                v_ext.focus_force()
                proveedores.SistemaProveedores(v_ext)
                self.v_prov.wait_window(v_ext)
                if not self.v_prov.winfo_exists() or not hasattr(self, 'cmb_cat_e') or not self.cmb_cat_e.winfo_exists():
                    return
                self.conn.commit()
                self.filtrar_proveedores_por_categoria()
            except Exception as e:
                try:
                    if self.v_prov.winfo_exists():
                        messagebox.showerror("Error", f"Falló la ejecución de Proveedores:\n{e}", parent=self.v_prov)
                except Exception:
                    pass

        ctk.CTkButton(f_prov_accion, text="[+] Crear Proveedor", width=120, command=abrir_sistema_proveedores).pack(side="left")

        ctk.CTkLabel(self.f_inputs, text="Cant.:", font=("Arial", 12, "bold")).grid(row=2, column=0, sticky="w", padx=(15, 2), pady=5)
        self.ent_cant = ctk.CTkEntry(self.f_inputs, width=60)
        self.ent_cant.grid(row=2, column=1, sticky="w", pady=5, padx=2)
        self.ent_cant.insert(0, "1")

        ctk.CTkLabel(self.f_inputs, text="P. Lista (S/.):", font=("Arial", 12, "bold")).grid(row=2, column=2, sticky="w", padx=(10, 2), pady=5)
        self.ent_p_lista = ctk.CTkEntry(self.f_inputs, width=100)
        self.ent_p_lista.grid(row=2, column=3, sticky="w", pady=5, padx=2)
        self.ent_p_lista.insert(0, "0.00")

        ctk.CTkLabel(self.f_inputs, text="P. Dscto (S/.):", font=("Arial", 12, "bold")).grid(row=3, column=0, sticky="w", pady=5, padx=(15, 2))
        self.ent_p_desc = ctk.CTkEntry(self.f_inputs, width=100)
        self.ent_p_desc.grid(row=3, column=1, sticky="w", pady=5, padx=2)
        self.ent_p_desc.insert(0, "0.00")

        ctk.CTkLabel(self.f_inputs, text="Tipo / Val. Ganancia:", font=("Arial", 12, "bold")).grid(row=3, column=2, sticky="w", pady=5, padx=(10, 2))
        f_gan_inline = ctk.CTkFrame(self.f_inputs, fg_color="transparent")
        f_gan_inline.grid(row=3, column=3, sticky="w", pady=5, padx=2)
        self.cmb_tipo_ganancia = ctk.CTkComboBox(f_gan_inline, values=["Monto Fijo", "Porcentaje (%)"], state="readonly", width=120)
        self.cmb_tipo_ganancia.pack(side="left", padx=(0, 2))
        self.cmb_tipo_ganancia.set("Monto Fijo")
        self.ent_val_g = ctk.CTkEntry(f_gan_inline, width=80)
        self.ent_val_g.pack(side="left")
        self.ent_val_g.insert(0, "0.00")

        ctk.CTkLabel(self.f_inputs, text="Tipo de Cambio ($):", font=("Arial", 12, "bold")).grid(row=3, column=5, sticky="e", padx=(15, 5), pady=5)
        self.var_tc_rastreador = tk.StringVar()
        self.var_tc_rastreador.trace_add("write", lambda *args: self._recalcular_con_retraso())
        self.ent_tc = ctk.CTkEntry(self.f_inputs, width=100, textvariable=self.var_tc_rastreador)
        self.ent_tc.grid(row=3, column=6, sticky="w", padx=(5, 15), pady=5)
        self.ent_tc.bind("<FocusOut>", lambda e: self.guardar_ajustes_globales_db())
        self.ent_tc.bind("<Return>", lambda e: self.guardar_ajustes_globales_db())

        ctk.CTkLabel(self.f_inputs, text="Forma de Pago (PDF):", font=("Arial", 12, "bold")).grid(row=4, column=5, sticky="e", padx=(15, 5), pady=(10, 5))
        self.ent_forma_pago = ctk.CTkEntry(self.f_inputs, width=320)
        self.ent_forma_pago.grid(row=4, column=6, sticky="w", padx=(5, 15), pady=(10, 5))
        self.ent_forma_pago.bind("<FocusOut>", lambda e: self.guardar_ajustes_globales_db())
        self.ent_forma_pago.bind("<Return>", lambda e: self.guardar_ajustes_globales_db())

        self.cargando_ventana = False

        f_notas_wrapper = ctk.CTkFrame(self.f_inputs, fg_color="transparent")
        f_notas_wrapper.grid(row=4, column=0, columnspan=5, rowspan=4, sticky="nw", pady=(12, 10), padx=15)
        
        f_nc = ctk.CTkFrame(f_notas_wrapper, fg_color="transparent")
        f_nc.pack(side="left", fill="both", expand=True, padx=(0, 10))
        f_header_nc = ctk.CTkFrame(f_nc, fg_color="transparent")
        f_header_nc.pack(fill="x", side="top", pady=(0, 2))
        ctk.CTkLabel(f_header_nc, text="Notas al Cliente (PDF/Excel):", font=("Arial", 11, "bold")).pack(side="left", anchor="w")
        self.txt_p_notes = ctk.CTkTextbox(f_nc, width=280, height=100, font=("Helvetica", 11), fg_color="#ffffff", text_color="#000000", border_width=1, border_color="#cccccc", corner_radius=5, wrap="word")
        f_estilos = crear_barra_formato(f_header_nc, self.txt_p_notes)
        f_estilos.pack(side="right", anchor="e")
        self.txt_p_notes.pack(fill="both", expand=True, side="top")
        
        f_ni = ctk.CTkFrame(f_notas_wrapper, fg_color="transparent")
        f_ni.pack(side="left", fill="both", expand=True, padx=(10, 0))
        ctk.CTkLabel(f_ni, text="Notas Internas (Solo Matriz):", font=("Arial", 11, "bold"), text_color="#D32F2F").pack(anchor="w", pady=(0, 2))
        self.txt_internal_notes = ctk.CTkTextbox(f_ni, width=280, height=100, font=("Helvetica", 11), fg_color="#FFFDE7", text_color="#000000", border_width=1, border_color="#FBC02D", corner_radius=5, wrap="word")
        self.txt_internal_notes.pack(fill="both", expand=True, side="top")

        f_totales_centro = ctk.CTkFrame(self.f_inputs, border_width=1, border_color="#cccccc", fg_color="#f9f9f9")
        f_totales_centro.grid(row=5, column=5, columnspan=2, rowspan=3, sticky="nsew", padx=(15, 15), pady=(5, 12))
        ctk.CTkLabel(f_totales_centro, text="Resumen Económico Contable", font=("Arial", 13, "bold"), text_color="#1f538d").pack(anchor="w", padx=15, pady=(10, 5))
        
        self.var_sin_fee = tk.BooleanVar(value=False)
        self.chk_sin_fee = ctk.CTkCheckBox(f_totales_centro, text="Exonerar Fee Producción (15%)", variable=self.var_sin_fee, command=self.evento_toggle_fee, text_color="#D32F2F", fg_color="#D32F2F", hover_color="#B71C1C")
        self.chk_sin_fee.pack(anchor="w", padx=15, pady=(0, 5))

        self.lbl_tot_sub = ctk.CTkLabel(f_totales_centro, text="Total Venta al Cliente: S/ 0.00", font=("Arial", 12, "bold"), text_color="#111111")
        self.lbl_tot_sub.pack(anchor="w", padx=15, pady=2)
        self.lbl_tot_igv = ctk.CTkLabel(f_totales_centro, text="15% Fee Producción: S/ 0.00", font=("Arial", 12, "bold"), text_color="#444444")
        self.lbl_tot_igv.pack(anchor="w", padx=15, pady=2)
        self.lbl_tot_gran = ctk.CTkLabel(f_totales_centro, text="Gran Total: S/ 0.00", font=("Arial", 14, "bold"), text_color="#e62060")
        self.lbl_tot_gran.pack(anchor="w", padx=15, pady=2)
        self.lbl_tot_usd = ctk.CTkLabel(f_totales_centro, text="Total Equivalente: $ 0.00 USD", font=("Arial", 12, "bold"), text_color="#222222")
        self.lbl_tot_usd.pack(anchor="w", padx=15, pady=(5, 2))

        # Ganancia total de la cotización:
        # (Venta + IGV) − (Compra + IGV) − Diferencial IGV − Renta
        # Renta = (Venta sin IGV − Detracción) × % Renta Mensual (Config. General)
        self.lbl_tot_costo = ctk.CTkLabel(f_totales_centro, text="Total Compra (sin IGV): S/ 0.00", font=("Arial", 12, "bold"), text_color="#555555")
        self.lbl_tot_costo.pack(anchor="w", padx=15, pady=2)
        self.lbl_igv_ventas = ctk.CTkLabel(f_totales_centro, text="IGV Ventas (18%): S/ 0.00", font=("Arial", 11), text_color="#7f8c8d")
        self.lbl_igv_ventas.pack(anchor="w", padx=15, pady=1)
        self.lbl_igv_compras = ctk.CTkLabel(f_totales_centro, text="IGV Compras (18%): S/ 0.00", font=("Arial", 11), text_color="#7f8c8d")
        self.lbl_igv_compras.pack(anchor="w", padx=15, pady=1)
        self.lbl_dif_igv = ctk.CTkLabel(f_totales_centro, text="Diferencial IGV (V−C): S/ 0.00", font=("Arial", 11), text_color="#34495e")
        self.lbl_dif_igv.pack(anchor="w", padx=15, pady=1)
        self.lbl_detraccion = ctk.CTkLabel(f_totales_centro, text="Detracción (12%): S/ 0.00", font=("Arial", 11), text_color="#8B4513")
        self.lbl_detraccion.pack(anchor="w", padx=15, pady=1)
        self.lbl_imp_renta = ctk.CTkLabel(f_totales_centro, text="Imp. Renta Mensual (1.5%): S/ 0.00", font=("Arial", 11), text_color="#8B4513")
        self.lbl_imp_renta.pack(anchor="w", padx=15, pady=1)
        self.lbl_tot_gan = ctk.CTkLabel(f_totales_centro, text="GANANCIA TOTAL: S/ 0.00", font=("Arial", 15, "bold"), text_color="#1e8449")
        self.lbl_tot_gan.pack(anchor="w", padx=15, pady=(6, 12))

        self.cargar_ajustes_globales()

        def disparar_exportacion_pdf_alberto():
            self.guardar_ajustes_globales_db()
            try:
                try:
                    import final_cotizaciones as motor_pdf
                    importlib.reload(motor_pdf) 
                except Exception:
                    import cotizaciones as motor_pdf
                    importlib.reload(motor_pdf)
                    
                conn_pdf = conectar_db(silencioso=True)
                if not conn_pdf: return
                try:
                    exito, mensaje_o_ruta = motor_pdf.generar_reporte_cotizacion_pdf(conn_pdf, self.codigo_cot)
                    if exito:
                        cache_sistema.invalidar()
                        registrar_auditoria(self.usuario_activo, "Cotizaciones", f"Exportó a PDF la Matriz Cotización N° {self.codigo_cot}")
                        messagebox.showinfo("Éxito de Exportación", f"¡Excelente!\nLa cotización oficial N° {self.codigo_cot} ha sido fabricada.\n\nArchivo guardado:\n{os.path.basename(mensaje_o_ruta)}", parent=self.v_prov)
                        try:
                            abrir_documento(mensaje_o_ruta)
                        except Exception: pass
                    else:
                        messagebox.showerror("Error de Creación", f"No se pudo maquetar el reporte contable:\n\n{mensaje_o_ruta}", parent=self.v_prov)
                finally:
                    liberar_conexion(conn_pdf)
            except Exception as e:
                messagebox.showerror("Error de Archivo", f"No se encontró el generador de PDF.\n\nDetalle: {str(e)}", parent=self.v_prov)

        btn_pdf = ctk.CTkButton(self.f_inputs, text="[ PDF ] Generar Cotización Oficial", font=("Arial", 14, "bold"), height=40, fg_color=COLOR_PRIMARIO, hover_color="#b71c1c", command=disparar_exportacion_pdf_alberto)
        btn_pdf.grid(row=8, column=0, columnspan=4, sticky="ew", padx=(15, 5), pady=(15, 10))
        
        btn_excel_int = ctk.CTkButton(self.f_inputs, text="[ EXCEL ] Exportar Matriz Interactiva", font=("Arial", 14, "bold"), height=40, fg_color="#27ae60", hover_color="#1e8449", command=self.exportar_matriz_excel_interactivo)
        btn_excel_int.grid(row=8, column=4, columnspan=4, sticky="ew", padx=(5, 15), pady=(15, 10))

        self.f_b_matriz = ctk.CTkFrame(self.v_prov, fg_color="transparent")
        self.f_b_matriz.pack(fill="x", padx=15, pady=5)
        ctk.CTkButton(self.f_b_matriz, text="<< Atrás", width=100, fg_color="#e0e0e0", text_color="black", hover_color="#c8c8c8", command=self.regresar_a_etapa2).pack(side="left", padx=(0, 5))
        self.btn_toggle_vista = ctk.CTkButton(self.f_b_matriz, text="[ - ] Mostrar Formulario", width=140, fg_color="#2980B9", hover_color="#1A5276", command=self.toggle_vista_matriz)
        self.btn_toggle_vista.pack(side="left", padx=5)
        ctk.CTkButton(self.f_b_matriz, text="[ Editar ] Costo/Margen", width=150, fg_color="#f39c12", hover_color="#e67e22", command=self.modificar_proveedor_matriz).pack(side="left", padx=5)
        ctk.CTkButton(self.f_b_matriz, text="[ X ] Retirar Ítem", width=120, fg_color="#D32F2F", hover_color="#B71C1C", command=self.retirar_proveedor_matriz).pack(side="left", padx=5)
        ctk.CTkButton(self.f_b_matriz, text="▲ Subir", width=70, fg_color="#e0e0e0", text_color="black", hover_color="#c8c8c8", command=lambda: self.mover_renglon_matriz("ARRIBA")).pack(side="left", padx=5)
        ctk.CTkButton(self.f_b_matriz, text="▼ Bajar", width=70, fg_color="#e0e0e0", text_color="black", hover_color="#c8c8c8", command=lambda: self.mover_renglon_matriz("ABAJO")).pack(side="left", padx=5)
        ctk.CTkButton(self.f_b_matriz, text="[ OK ] Finalizar Matriz", width=150, command=self._cerrar_ventana).pack(side="right", padx=5)
        ctk.CTkButton(self.f_b_matriz, text="[ + ] Asignar a Matriz", width=150, fg_color="#228B22", hover_color="#1E761E", command=self.asociar_proveedor_a_matriz).pack(side="right", padx=5)

        self.f_grid = ctk.CTkFrame(self.v_prov, corner_radius=10)
        self.f_grid.pack(fill="both", expand=True, padx=15, pady=(5, 10))
        ctk.CTkLabel(self.f_grid, text="Matriz Comparativa y Margen Final de Venta", font=("Arial", 14, "bold"), text_color="#1f538d").pack(anchor="w", padx=15, pady=10)
        
        f_headers = ctk.CTkFrame(self.f_grid, fg_color="#e0e0e0", corner_radius=5)
        f_headers.pack(fill="x", padx=10, pady=(0, 5))
        
        anchos_encabezado = [("ID", 30), ("Categoría", 130), ("Proveedor", 140), ("Cant.", 40), ("P. Lista", 70), ("Dscto", 50), ("P. Venta", 80), ("Ganancia", 85)]
        for text, w in anchos_encabezado:
            ctk.CTkLabel(f_headers, text=text, font=("Arial", 11, "bold"), width=w, anchor="center").pack(side="left", padx=2, pady=5)
        
        f_h_notas = ctk.CTkFrame(f_headers, fg_color="transparent")
        f_h_notas.pack(side="left", fill="x", expand=True, padx=2)
        ctk.CTkLabel(f_h_notas, text="Notas Cliente", font=("Arial", 11, "bold"), anchor="center").pack(side="left", fill="x", expand=True, padx=4)
        ctk.CTkLabel(f_h_notas, text="Notas Internas", font=("Arial", 11, "bold"), text_color="#D32F2F", anchor="center").pack(side="left", fill="x", expand=True, padx=4)
        
        self.f_rows_dinamicas = ctk.CTkScrollableFrame(self.f_grid, fg_color="transparent")
        self.f_rows_dinamicas.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.filtrar_proveedores_por_categoria()
        self.cargar_grid_proveedores()
        
        self.matriz_expandida = False

    def evento_toggle_fee(self):
        self.guardar_ajustes_globales_db()
        self._recalcular_con_retraso()

    def exportar_matriz_excel_interactivo(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Protection, Border, Side
            from openpyxl.drawing.image import Image as OpenpyxlImage
        except ImportError:
            messagebox.showerror("Librería faltante", "Falta la librería openpyxl o Pillow.\n\nAbre tu terminal y ejecuta:\npip install openpyxl Pillow", parent=self.v_prov)
            return

        if not self.conn: return

        try:
            config_data = {}
            try:
                with open(RUTA_CONFIG, "r", encoding="utf-8") as f:
                    config_data = json.load(f)
            except Exception:
                pass
                
            color_primario = config_data.get("color_primario", COLOR_PRIMARIO).replace("#", "").upper()
            if len(color_primario) != 6: color_primario = "EB337A"
            logo_ruta = config_data.get("ruta_logo_cotizacion", "")

            c = self.conn.cursor()
            
            c.execute("SELECT sin_fee FROM cotizaciones WHERE codigo_cotizacion = %s", (self.codigo_cot,))
            res_fee = c.fetchone()
            sin_fee_db = bool(res_fee[0]) if res_fee and res_fee[0] is not None else False
            
            c.execute("SELECT cantidad, notes_negociacion, precio_final_venta FROM cotizacion_proveedores WHERE codigo_cotizacion = %s ORDER BY id ASC", (self.codigo_cot,))
            registros = c.fetchall()

            if not registros:
                messagebox.showwarning("Sin Datos", "No hay costos asignados para exportar.", parent=self.v_prov)
                return

            ruta = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=f"Cotizacion_Interactiva_{self.codigo_cot}.xlsx", filetypes=[("Excel", "*.xlsx")], parent=self.v_prov)
            if not ruta: return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cotización"
            ws.sheet_view.showGridLines = False 

            header_fill = PatternFill(start_color=color_primario, end_color=color_primario, fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            unlocked = Protection(locked=False)
            fill_editable = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
            border_thin = Border(left=Side(style='thin', color="DDDDDD"), 
                                 right=Side(style='thin', color="DDDDDD"), 
                                 top=Side(style='thin', color="DDDDDD"), 
                                 bottom=Side(style='thin', color="DDDDDD"))

            if logo_ruta and os.path.exists(logo_ruta):
                try:
                    img = OpenpyxlImage(logo_ruta)
                    aspect_ratio = img.width / img.height
                    nuevo_ancho = 830
                    img.width = nuevo_ancho
                    img.height = nuevo_ancho / aspect_ratio
                    ws.add_image(img, 'A1')
                    ws.row_dimensions[1].height = (img.height * 0.75) + 5
                except Exception:
                    ws.row_dimensions[1].height = 20
            else:
                ws.row_dimensions[1].height = 20

            ws.merge_cells('A2:D2')
            ws['A2'] = f"COTIZACIÓN N° {self.codigo_cot}"
            ws['A2'].font = Font(size=14, bold=True, color=color_primario)
            ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
            
            ws.merge_cells('A3:D3')
            ws['A3'] = f"Cliente: {self.empresa}"
            ws['A3'].font = Font(size=12, bold=True)
            ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
            
            ws.merge_cells('A4:D4')
            ws['A4'] = f"Evento / Servicio: {self.evento}"
            ws['A4'].font = Font(size=12)
            ws['A4'].alignment = Alignment(horizontal="center", vertical="center")
            
            ws.merge_cells('A5:D5')
            ws['A5'] = "Instrucciones: Solo puede modificar las celdas celestes (Cantidad y P. Unitario). Los totales se calcularán automáticamente."
            ws['A5'].font = Font(italic=True, color="555555")
            ws['A5'].alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[2].height = 25
            ws.row_dimensions[3].height = 20
            ws.row_dimensions[4].height = 20
            ws.row_dimensions[5].height = 20
            ws.row_dimensions[6].height = 10 

            headers = ["Cantidad", "Notas Adicionales", "P. Unitario", "Subtotal Fila"]
            row_idx = 7
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_thin

            row_idx = 8
            for r in registros:
                cant = int(r[0]) if r[0] else 1
                notas = texto_plano_sin_marcado(str(r[1])) if r[1] else ""
                precio_tot = float(r[2])
                p_unit = precio_tot / cant if cant else 0.0

                c_cant = ws.cell(row=row_idx, column=1, value=cant)
                c_cant.protection = unlocked
                c_cant.fill = fill_editable
                c_cant.alignment = Alignment(horizontal="center", vertical="center")
                c_cant.border = border_thin

                c_notas = ws.cell(row=row_idx, column=2, value=notas)
                c_notas.alignment = Alignment(wrap_text=True, vertical="center")
                c_notas.border = border_thin

                c_price = ws.cell(row=row_idx, column=3, value=p_unit)
                c_price.protection = unlocked
                c_price.fill = fill_editable
                c_price.number_format = '"S/." #,##0.00'
                c_price.alignment = Alignment(vertical="center")
                c_price.border = border_thin

                c_subtot = ws.cell(row=row_idx, column=4, value=f"=A{row_idx}*C{row_idx}")
                c_subtot.number_format = '"S/." #,##0.00'
                c_subtot.font = Font(bold=True)
                c_subtot.alignment = Alignment(vertical="center")
                c_subtot.border = border_thin

                row_idx += 1

            row_idx += 1
            subtotal_row = row_idx
            ws.cell(row=row_idx, column=3, value="SUBTOTAL:").font = Font(bold=True)
            ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right")
            ws.cell(row=row_idx, column=4, value=f"=SUM(D8:D{row_idx-2})").number_format = '"S/." #,##0.00'
            ws.cell(row=row_idx, column=4).font = Font(bold=True)

            if not sin_fee_db:
                row_idx += 1
                fee_row = row_idx
                ws.cell(row=row_idx, column=3, value="FEE (15%):").font = Font(bold=True)
                ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right")
                ws.cell(row=row_idx, column=4, value=f"=D{subtotal_row}*0.15").number_format = '"S/." #,##0.00'
                ws.cell(row=row_idx, column=4).font = Font(bold=True)

            row_idx += 1
            ws.cell(row=row_idx, column=3, value="GRAN TOTAL:").font = Font(bold=True, color=color_primario)
            ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right")
            
            if not sin_fee_db:
                c_gran = ws.cell(row=row_idx, column=4, value=f"=D{subtotal_row}+D{fee_row}")
            else:
                c_gran = ws.cell(row=row_idx, column=4, value=f"=D{subtotal_row}")
                
            c_gran.number_format = '"S/." #,##0.00'
            c_gran.font = Font(bold=True, color=color_primario)

            ws.column_dimensions['A'].width = 15  
            ws.column_dimensions['B'].width = 65  
            ws.column_dimensions['C'].width = 18  
            ws.column_dimensions['D'].width = 20  

            ws.protection.sheet = True
            ws.protection.password = "blackcube2026" 

            try:
                wb.save(ruta)
            except PermissionError:
                messagebox.showwarning("Archivo Abierto", f"No se puede guardar porque el archivo Excel ya está abierto.\n\nPor favor, cierra:\n{os.path.basename(ruta)}\n\ny vuelve a intentarlo.", parent=self.v_prov)
                return

            messagebox.showinfo("Éxito", f"Excel Interactivo generado en:\n{ruta}\n\nEl archivo tiene diseño corporativo y está protegido.\n\nEl cliente solo podrá modificar las celdas celestes (Cantidades y P. Unitario).", parent=self.v_prov)
            abrir_documento(ruta)
            registrar_auditoria(self.usuario_activo, "Cotizaciones", f"Exportó a Excel Interactivo la Matriz Cotización N° {self.codigo_cot}")
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar a Excel:\n{str(e)}", parent=self.v_prov)

    # =======================================================
    # TIPO DE CAMBIO EN VIVO (EN SEGUNDO PLANO, NO CONGELA)
    # =======================================================
    def obtener_tipo_cambio_en_vivo(self):
        try:
            url = "https://api.apis.net.pe/v1/tipo-cambio-sunat"
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=3) as response:
                data = json.loads(response.read().decode())
                return float(data.get("venta", 3.75))
        except Exception:
            try:
                url = "https://open.er-api.com/v6/latest/USD"
                req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=3) as response:
                    data = json.loads(response.read().decode())
                    return float(data["rates"]["PEN"])
            except Exception:
                return None

    def _cargar_tipo_cambio_en_vivo(self):
        def tarea():
            tc = self.obtener_tipo_cambio_en_vivo()

            def aplicar():
                try:
                    if not self.v_prov.winfo_exists():
                        return
                except Exception:
                    return
                if tc:
                    self.ent_tc.delete(0, tk.END)
                    self.ent_tc.insert(0, f"{tc:.3f}")
                self.guardar_ajustes_globales_db()

            self.root.after(0, aplicar)

        threading.Thread(target=tarea, daemon=True).start()

    # =======================================================
    # AJUSTES GLOBALES (TC Y FORMA DE PAGO)
    # =======================================================
    def cargar_ajustes_globales(self):
        try:
            c = self.conn.cursor()
            c.execute("SELECT tipo_cambio, forma_pago, sin_fee FROM cotizaciones WHERE codigo_cotizacion = %s", (self.codigo_cot,))
            res = c.fetchone()
            if res:
                self.ent_tc.delete(0, tk.END)
                self.ent_tc.insert(0, str(res[0]) if res[0] and float(res[0]) > 0 else "3.750")
                self.ent_forma_pago.delete(0, tk.END)
                self.ent_forma_pago.insert(0, str(res[1]) if res[1] else "50% adelantado, 50% a 30 días de la primera factura.")
                self.var_sin_fee.set(bool(res[2]) if res[2] is not None else False)
        except Exception:
            try:
                self.ent_tc.insert(0, "3.750")
                self.ent_forma_pago.insert(0, "50% adelantado, 50% a 30 días de la primera factura.")
            except Exception:
                pass
        self._cargar_tipo_cambio_en_vivo()

    def guardar_ajustes_globales_db(self):
        if not self.conn:
            return
        try:
            c = self.conn.cursor()
            val_tc = 3.75
            if self.ent_tc.get().strip():
                try:
                    val_tc = float(self.ent_tc.get())
                except ValueError:
                    pass
            val_forma_pago = self.ent_forma_pago.get().strip()
            val_sin_fee = self.var_sin_fee.get()
            c.execute("UPDATE cotizaciones SET tipo_cambio = %s, forma_pago = %s, sin_fee = %s WHERE codigo_cotizacion = %s", (val_tc, val_forma_pago, val_sin_fee, self.codigo_cot))
            self.conn.commit()
        except Exception:
            pass

    # =======================================================
    # TOTALES CON RETRASO INTELIGENTE + HILO (NO CONGELA)
    # =======================================================
    def _recalcular_con_retraso(self, *args):
        if hasattr(self, "_recalc_job"):
            try:
                self.root.after_cancel(self._recalc_job)
            except Exception:
                pass
        self._recalc_job = self.root.after(300, self.actualizar_bloque_totales_pantalla)

    def actualizar_bloque_totales_pantalla(self):
        if getattr(self, "cargando_ventana", False):
            return
        try:
            tc = float(self.ent_tc.get())
        except Exception:
            tc = 3.75
        codigo = self.codigo_cot

        def tarea():
            subtotal = 0.0
            costo_total = 0.0
            conn = conectar_db(silencioso=True)
            if conn:
                try:
                    c = conn.cursor()
                    c.execute("SELECT precio_lista, precio_descuento, precio_final_venta, cantidad FROM cotizacion_proveedores WHERE codigo_cotizacion = %s", (codigo,))
                    for r in c.fetchall():
                        if not r:
                            continue
                        pl_r, pd_r, pf_r, cant_r = r[0], r[1], r[2], (r[3] if len(r) > 3 and r[3] else 1)
                        if pf_r:
                            subtotal += float(pf_r)
                        costo_u = float(pd_r) if pd_r and float(pd_r) > 0 else float(pl_r or 0)
                        costo_total += costo_u * float(cant_r or 1)
                except Exception:
                    subtotal = 0.0
                    costo_total = 0.0
                finally:
                    liberar_conexion(conn)
            self.root.after(0, lambda s=subtotal, t=tc, ct=costo_total: self._pintar_totales(s, t, ct))

        threading.Thread(target=tarea, daemon=True).start()

    def _obtener_porcentaje_config(self, clave, defecto):
        """Lee un porcentaje de la Configuración General (control_general.py).
        Claves: igv_porcentaje, retencion_porcentaje, detraccion_porcentaje,
        renta_mensual_porcentaje, renta_anual_porcentaje."""
        try:
            with open(RUTA_CONFIG, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            val = cfg.get(clave)
            return float(val) if val not in (None, "") else float(defecto)
        except Exception:
            return float(defecto)

    def _pintar_totales(self, subtotal, tc, costo_total=0.0):
        try:
            if not self.v_prov.winfo_exists():
                return
        except Exception:
            return
            
        aplica_fee = not self.var_sin_fee.get()
        fee = subtotal * 0.15 if aplica_fee else 0.0
        
        try:
            tc_val = float(self.ent_tc.get())
        except Exception:
            tc_val = tc if tc else 3.75
            
        self.lbl_tot_sub.configure(text=f"Total Venta al Cliente: S/ {subtotal:,.2f}")
        
        if aplica_fee:
            self.lbl_tot_igv.configure(text=f"15% Fee Producción: S/ {fee:,.2f}")
        else:
            self.lbl_tot_igv.configure(text=f"Fee Producción: S/ 0.00 (Exonerado)")
            
        self.lbl_tot_gran.configure(text=f"Gran Total: S/ {subtotal + fee:,.2f}")
        self.lbl_tot_usd.configure(text=f"Total Equivalente: $ {(subtotal + fee) / tc_val:,.2f} USD")

        # ── GANANCIA TOTAL DE LA COTIZACIÓN ─────────────────────────────
        # Fórmula: (Venta + IGV) − (Compra + IGV) − Diferencial IGV − Renta
        # Renta = (Venta sin IGV − Detracción) × % Renta Mensual
        # Los porcentajes salen de la Configuración General (control_general.py).
        igv_pct = self._obtener_porcentaje_config("igv_porcentaje", 18)
        detraccion_pct = self._obtener_porcentaje_config("detraccion_porcentaje", 12)
        renta_pct = self._obtener_porcentaje_config("renta_mensual_porcentaje", 1.5)

        venta_sin_igv = subtotal
        compra_sin_igv = costo_total
        igv_ventas = venta_sin_igv * igv_pct / 100.0
        igv_compras = compra_sin_igv * igv_pct / 100.0
        dif_igv = igv_ventas - igv_compras
        detraccion = venta_sin_igv * detraccion_pct / 100.0
        impuesto_renta = (venta_sin_igv - detraccion) * renta_pct / 100.0
        ganancia_neta = (venta_sin_igv + igv_ventas) - (compra_sin_igv + igv_compras) - dif_igv - impuesto_renta

        self.lbl_tot_costo.configure(text=f"Total Compra (sin IGV): S/ {compra_sin_igv:,.2f}")
        self.lbl_igv_ventas.configure(text=f"IGV Ventas ({igv_pct:g}%): S/ {igv_ventas:,.2f}")
        self.lbl_igv_compras.configure(text=f"IGV Compras ({igv_pct:g}%): S/ {igv_compras:,.2f}")
        self.lbl_dif_igv.configure(text=f"Diferencial IGV (V−C): S/ {dif_igv:,.2f}")
        self.lbl_detraccion.configure(text=f"Detracción ({detraccion_pct:g}%): S/ {detraccion:,.2f}")
        self.lbl_imp_renta.configure(text=f"Imp. Renta Mensual ({renta_pct:g}%): S/ {impuesto_renta:,.2f}")
        self.lbl_tot_gan.configure(text=f"GANANCIA TOTAL: S/ {ganancia_neta:,.2f}")

    # =======================================================
    # VISTAS Y NAVEGACIÓN
    # =======================================================
    def toggle_vista_matriz(self):
        if not self.matriz_expandida:
            self.f_inputs.pack_forget()
            self.btn_toggle_vista.configure(text="[ - ] Mostrar Formulario", fg_color="#2980B9", hover_color="#1A5276")
            self.matriz_expandida = True
        else:
            self.f_inputs.pack(fill="x", padx=15, pady=5, after=self.f_info)
            self.btn_toggle_vista.configure(text="[ + ] Pantalla Completa", fg_color="#8E44AD", hover_color="#732D91")
            self.matriz_expandida = False

    def _liberar_conn(self):
        try:
            if self.conn:
                liberar_conexion(self.conn)
        except Exception:
            pass
        self.conn = None

    def _cerrar_ventana(self):
        self._liberar_conn()
        try:
            self.v_prov.destroy()
        except Exception:
            pass
        if self.callback_on_close:
            self.callback_on_close()

    def regresar_a_etapa2(self):
        self._liberar_conn()
        self.v_prov.destroy()
        if hasattr(self.parent_ventana, 'abrir_ventana_editar'):
            self.parent_ventana.abrir_ventana_editar(codigo_directo=self.codigo_cot)
        elif self.callback_on_close:
            self.callback_on_close()

    def refrescar_combobox_descarte(self):
        if not self.conn:
            return
        try:
            c = self.conn.cursor()
            c.execute("SELECT categoria_suministro FROM cotizacion_detalles WHERE codigo_cotizacion = %s", (self.codigo_cot,))
            nuevas = []
            for r in c.fetchall():
                cat = str(r[0]).replace("('", "").replace("',)", "").replace("',", "").strip("() '\",")
                if cat and cat not in nuevas:
                    nuevas.append(cat)
            if not nuevas:
                nuevas = ["No hay categorias disponibles"]
            sel_actual = self.cmb_cat_e.get()
            self.cmb_cat_e.configure(values=nuevas)
            if sel_actual in nuevas:
                self.cmb_cat_e.set(sel_actual)
            else:
                self.cmb_cat_e.set(nuevas[0])
        except Exception:
            pass

    def filtrar_proveedores_por_categoria(self, choice=None):
        cat_sel = str(self.cmb_cat_e.get()).strip()
        self.cmb_p_list.set("Cargando...")
        
        def proceso_pesado():
            clave_cache = "lista_proveedores_texto_completo_v2"
            lista = cache_sistema.obtener(clave_cache)
            
            if not lista:
                conn = conectar_db(silencioso=True)
                if conn:
                    try:
                        cursor = conn.cursor()
                        cursor.execute("SELECT nombre, t::text FROM proveedores t ORDER BY nombre ASC")
                        lista = [(str(r[0]).strip(), str(r[1]).strip() if r[1] else "") for r in cursor.fetchall() if r[0]]
                        cache_sistema.guardar(clave_cache, lista)
                    except: pass
                    finally: liberar_conexion(conn)

            if not lista: lista = []

            def normalizar(t):
                if not t: return ""
                res = _NORM_CACHE_PROVEEDORES.get(t)
                if res is None:
                    res = ''.join(c for c in unicodedata.normalize('NFD', str(t).lower().strip()) if unicodedata.category(c) != 'Mn')
                    _NORM_CACHE_PROVEEDORES[t] = res
                return res

            cat_norm = normalizar(cat_sel)
            palabras_clave = [p for p in cat_norm.split() if len(p) >= 3]
            if not palabras_clave: palabras_clave = [cat_norm]

            provs_filtrados = []
            for n, c in lista:
                c_norm = normalizar(c)
                match = False
                if cat_norm in c_norm:
                    match = True
                else:
                    for p in palabras_clave:
                        if p in c_norm:
                            match = True
                            break
                if match:
                    provs_filtrados.append(n)

            provs_filtrados = list(dict.fromkeys(provs_filtrados))
            if provs_filtrados:
                lista_final = ["--- Seleccione Proveedor ---"] + provs_filtrados
            else:
                lista_final = ["--- Seleccione (No hay del rubro) ---"] + [n for n, c in lista]

            if hasattr(self, 'root') and self.v_prov.winfo_exists():
                self.root.after(0, lambda: self._actualizar_ui_combo(lista_final))

        threading.Thread(target=proceso_pesado, daemon=True).start()

    def _actualizar_ui_combo(self, lista_final):
        if not getattr(self, 'v_prov', None) or not self.v_prov.winfo_exists(): return
        prov_actual = self.cmb_p_list.get()
        self.cmb_p_list.configure(values=lista_final)
        if prov_actual in lista_final and "Seleccione" not in prov_actual:
            self.cmb_p_list.set(prov_actual)
        else:
            self.cmb_p_list.set(lista_final[0])

    # =======================================================
    # OPERACIONES DE MATRIZ (CON BITÁCORA)
    # =======================================================
    def asociar_proveedor_a_matriz(self):
        if not self.conn:
            return
        cat, prov = self.cmb_cat_e.get().strip(), self.cmb_p_list.get().strip()
        if prov in ["Seleccione un proveedor", "Haga clic en Cargar Proveedores", "", "--- Sin proveedores específicos, mostrando todos ---", "--- Seleccione Proveedor ---", "--- Seleccione (No hay del rubro) ---"]:
            messagebox.showwarning("Atención", "Por favor despliegue la lista y seleccione un proveedor válido.", parent=self.v_prov)
            return
            
        try:
            cant = int(self.ent_cant.get().strip())
            dias_cred = 0 
            pl, pd, vg = float(self.ent_p_lista.get()), float(self.ent_p_desc.get()), float(self.ent_val_g.get())
        except ValueError:
            messagebox.showwarning("Error numérico", "Importes o cantidades inválidas.", parent=self.v_prov)
            return
            
        if pd > pl:
            messagebox.showwarning("Alerta", "El descuento no puede superar al precio de lista.", parent=self.v_prov)
            return
            
        p_unid = pl + vg if self.cmb_tipo_ganancia.get() == "Monto Fijo" else pl * (1 + (vg / 100.0))
        p_final_venta = p_unid * cant
        t_ganancia_db = "Monto Fijo" if self.cmb_tipo_ganancia.get() == "Monto Fijo" else "Porcentaje"
        
        notes = extraer_texto_con_formato(self.txt_p_notes).strip()
        internal_notes = self.txt_internal_notes.get("1.0", tk.END).strip()
        
        c = self.conn.cursor()
        try:
            c.execute("""
                INSERT INTO cotizacion_proveedores 
                (codigo_cotizacion, categoria_suministro, proveedor_nombre, precio_lista, precio_descuento, tipo_ganancia, valor_ganancia, precio_final_venta, notes_negociacion, notas_internas, cantidad, dias_credito) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (self.codigo_cot, cat, prov, pl, pd, t_ganancia_db, vg, p_final_venta, notes, internal_notes, cant, dias_cred))
            self.conn.commit()
            cache_sistema.invalidar()
            registrar_auditoria(self.usuario_activo, "Cotizaciones", f"Agregó proveedor '{prov}' a Cotización N° {self.codigo_cot}")
            self.actualizar_bloque_totales_pantalla()
            self.ent_cant.delete(0, tk.END)
            self.ent_cant.insert(0, "1")
            self.ent_p_lista.delete(0, tk.END)
            self.ent_p_lista.insert(0, "0.00")
            self.ent_p_desc.delete(0, tk.END)
            self.ent_p_desc.insert(0, "0.00")
            self.ent_val_g.delete(0, tk.END)
            self.ent_val_g.insert(0, "0.00")
            self.txt_p_notes.delete("1.0", tk.END)
            self.txt_internal_notes.delete("1.0", tk.END)
            self.cargar_grid_proveedores()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo asignar al proveedor:\n{e}", parent=self.v_prov)

    def retirar_proveedor_matriz(self):
        if not self.conn:
            return
        if not self.fila_matriz_seleccionada:
            messagebox.showwarning("Advertencia", "Seleccione una fila de la matriz primero.", parent=self.v_prov)
            return
        if messagebox.askyesno("Confirmar", "¿Retirar al proveedor seleccionado?", parent=self.v_prov):
            try:
                c = self.conn.cursor()
                c.execute("DELETE FROM cotizacion_proveedores WHERE id=%s", (self.fila_matriz_seleccionada[0],))
                self.conn.commit()
                cache_sistema.invalidar()
                registrar_auditoria(self.usuario_activo, "Cotizaciones", f"Retiró ítem de la Cotización N° {self.codigo_cot}")
            except Exception as e:
                self.conn.rollback()
                messagebox.showerror("Error", f"No se pudo eliminar:\n{e}", parent=self.v_prov)
                return
            self.cargar_grid_proveedores()

    def modificar_proveedor_matriz(self):
        if not self.conn:
            return
        if not self.fila_matriz_seleccionada:
            messagebox.showwarning("Requerido", "Seleccione una fila primero.", parent=self.v_prov)
            return
        id_mat = self.fila_matriz_seleccionada[0]
        c = self.conn.cursor()
        c.execute("SELECT precio_lista, precio_descuento, tipo_ganancia, valor_ganancia, notes_negociacion, notas_internas, cantidad FROM cotizacion_proveedores WHERE id=%s", (id_mat,))
        datos = c.fetchone()
        if not datos:
            return
        v_m = ctk.CTkToplevel(self.v_prov)
        v_m.title("Modificar Costos")
        v_m.geometry("550x720")
        v_m.grab_set()
        f_m = ctk.CTkFrame(v_m, corner_radius=10)
        f_m.pack(fill="both", expand=True, padx=15, pady=15)
        
        ctk.CTkLabel(f_m, text="Cantidad:", font=("Arial", 12, "bold")).pack(anchor="w", padx=10, pady=(5, 2))
        ent_m_cant = ctk.CTkEntry(f_m, width=100)
        ent_m_cant.pack(anchor="w", padx=10, pady=2)
        ent_m_cant.insert(0, str(datos[6] if len(datos) > 6 and datos[6] else 1))
        
        ctk.CTkLabel(f_m, text="Precio Lista (S/.):", font=("Arial", 12, "bold")).pack(anchor="w", padx=10, pady=(5, 2))
        ent_m_lista = ctk.CTkEntry(f_m, width=200)
        ent_m_lista.pack(anchor="w", padx=10, pady=2)
        ent_m_lista.insert(0, f"{datos[0]:.2f}")
        
        ctk.CTkLabel(f_m, text="Precio con Descuento (S/.):", font=("Arial", 12, "bold")).pack(anchor="w", padx=10, pady=(5, 2))
        ent_m_desc = ctk.CTkEntry(f_m, width=200)
        ent_m_desc.pack(anchor="w", padx=10, pady=2)
        ent_m_desc.insert(0, f"{datos[1]:.2f}")
        
        ctk.CTkLabel(f_m, text="Tipo de Ganancia:", font=("Arial", 12, "bold")).pack(anchor="w", padx=10, pady=(5, 2))
        cmb_m_tipo = ctk.CTkComboBox(f_m, values=["Monto Fijo", "Porcentaje (%)"], state="readonly", width=200)
        cmb_m_tipo.pack(anchor="w", padx=10, pady=2)
        cmb_m_tipo.set(datos[2])
        
        ctk.CTkLabel(f_m, text="Valor de Ganancia:", font=("Arial", 12, "bold")).pack(anchor="w", padx=10, pady=(5, 2))
        ent_m_val = ctk.CTkEntry(f_m, width=200)
        ent_m_val.pack(anchor="w", padx=10, pady=2)
        ent_m_val.insert(0, f"{datos[3]:.2f}")
        
        f_m_header = ctk.CTkFrame(f_m, fg_color="transparent")
        f_m_header.pack(fill="x", pady=(10, 2), padx=10)
        ctk.CTkLabel(f_m_header, text="Notas al Cliente:", font=("Arial", 12, "bold")).pack(side="left")
        
        txt_m_notas = ctk.CTkTextbox(f_m, width=500, height=80, font=("Helvetica", 11), fg_color="#ffffff", text_color="#000000", border_width=1, border_color="#cccccc", corner_radius=5, wrap="word")
        f_barra = crear_barra_formato(f_m_header, txt_m_notas)
        f_barra.pack(side="right")
        txt_m_notas.pack(fill="x", padx=10, pady=2)
        insertar_texto_formateado(txt_m_notas, str(datos[4]) if datos[4] else "")

        ctk.CTkLabel(f_m, text="Notas Internas (Solo Matriz):", font=("Arial", 12, "bold"), text_color="#D32F2F").pack(anchor="w", padx=10, pady=(10, 2))
        txt_m_notas_internas = ctk.CTkTextbox(f_m, width=500, height=80, font=("Helvetica", 11), fg_color="#FFFDE7", text_color="#000000", border_width=1, border_color="#FBC02D", corner_radius=5, wrap="word")
        txt_m_notas_internas.pack(fill="x", padx=10, pady=2)
        txt_m_notas_internas.insert("1.0", str(datos[5]) if datos[5] else "")

        def ejecutar_update_matriz():
            try:
                mc = int(ent_m_cant.get().strip())
                m_dcred = 0 
                ml, md, mv = float(ent_m_lista.get()), float(ent_m_desc.get()), float(ent_m_val.get())
            except ValueError:
                messagebox.showwarning("Error numérico", "Valores inválidos.", parent=v_m)
                return
            p_unid = ml + mv if cmb_m_tipo.get() == "Monto Fijo" else ml * (1 + (mv / 100))
            p_final = p_unid * mc
            notes_update = extraer_texto_con_formato(txt_m_notas).strip()
            internal_notes_update = txt_m_notas_internas.get("1.0", tk.END).strip()
            
            c_upd = self.conn.cursor()
            c_upd.execute("UPDATE cotizacion_proveedores SET precio_lista=%s, precio_descuento=%s, tipo_ganancia=%s, valor_ganancia=%s, precio_final_venta=%s, notes_negociacion=%s, notas_internas=%s, cantidad=%s, dias_credito=%s WHERE id=%s", (ml, md, cmb_m_tipo.get(), mv, p_final, notes_update, internal_notes_update, mc, m_dcred, id_mat))
            self.conn.commit()
            cache_sistema.invalidar()
            registrar_auditoria(self.usuario_activo, "Cotizaciones", f"Modificó márgenes en ítem de Cotización N° {self.codigo_cot}")
            v_m.destroy()
            self.cargar_grid_proveedores()

        ctk.CTkButton(f_m, text="[ Guardar Cambios ]", command=ejecutar_update_matriz).pack(pady=15)

    # =======================================================
    # LÓGICA DE INTERCAMBIO (DRAG & DROP Y BOTONES)
    # =======================================================
    def _intercambiar_datos(self, id1, id2, cursor):
        cols = "categoria_suministro, proveedor_nombre, precio_lista, precio_descuento, tipo_ganancia, valor_ganancia, precio_final_venta, notes_negociacion, notas_internas, cantidad, dias_credito"
        cursor.execute(f"SELECT {cols} FROM cotizacion_proveedores WHERE id = %s", (id1,))
        d1 = cursor.fetchone()
        cursor.execute(f"SELECT {cols} FROM cotizacion_proveedores WHERE id = %s", (id2,))
        d2 = cursor.fetchone()
        if d1 and d2:
            query = f"UPDATE cotizacion_proveedores SET categoria_suministro=%s, proveedor_nombre=%s, precio_lista=%s, precio_descuento=%s, tipo_ganancia=%s, valor_ganancia=%s, precio_final_venta=%s, notes_negociacion=%s, notas_internas=%s, cantidad=%s, dias_credito=%s WHERE id=%s"
            cursor.execute(query, d2 + (id1,))
            cursor.execute(query, d1 + (id2,))

    def _reordenar_items(self, start_idx, target_idx):
        if not self.conn: return
        c = self.conn.cursor()
        try:
            if start_idx < target_idx:
                for i in range(start_idx, target_idx):
                    id1 = self.lista_widgets_filas[i].data_pack[0]
                    id2 = self.lista_widgets_filas[i+1].data_pack[0]
                    self._intercambiar_datos(id1, id2, c)
            else:
                for i in range(start_idx, target_idx, -1):
                    id1 = self.lista_widgets_filas[i].data_pack[0]
                    id2 = self.lista_widgets_filas[i-1].data_pack[0]
                    self._intercambiar_datos(id1, id2, c)
            self.conn.commit()
            cache_sistema.invalidar()
            id_dest = self.lista_widgets_filas[target_idx].data_pack[0]
            self.cargar_grid_proveedores(id_a_seleccionar=id_dest)
        except Exception as e:
            self.conn.rollback()
            print("Error Drag&Drop:", e)
            self.cargar_grid_proveedores()

    def _iniciar_arrastre(self, event, idx):
        self._drag_start_index = idx
        if idx < len(self.lista_widgets_filas):
            f_row = self.lista_widgets_filas[idx]
            f_row.configure(border_color=COLOR_PRIMARIO, border_width=2)

    def _en_arrastre(self, event):
        self.v_prov.config(cursor="fleur")

    def _soltar_arrastre(self, event):
        self.v_prov.config(cursor="")
        if not hasattr(self, '_drag_start_index'): return
        start_idx = self._drag_start_index
        del self._drag_start_index
        
        y_mouse = event.y_root
        target_idx = -1
        
        for i, row in enumerate(self.lista_widgets_filas):
            y_row = row.winfo_rooty()
            h_row = row.winfo_height()
            if y_row <= y_mouse <= y_row + h_row:
                target_idx = i
                break
                
        if target_idx == -1 and self.lista_widgets_filas:
            if y_mouse < self.lista_widgets_filas[0].winfo_rooty():
                target_idx = 0
            elif y_mouse > self.lista_widgets_filas[-1].winfo_rooty() + self.lista_widgets_filas[-1].winfo_height():
                target_idx = len(self.lista_widgets_filas) - 1
                
        if target_idx != -1 and start_idx != target_idx:
            self._reordenar_items(start_idx, target_idx)
        else:
            if start_idx < len(self.lista_widgets_filas):
                f_row = self.lista_widgets_filas[start_idx]
                f_row.configure(border_color="#e0e0e0", border_width=1)

    # =======================================================
    # PINTADO DE LA MATRIZ (GRILLA RESTRINGIDA CON WRAP)
    # =======================================================
    # Nota scroll: el scroll de la matriz y de las cajas de notas se maneja de
    # forma global e inteligente en scroll_utils.py (la rueda desplaza primero
    # el contenido de la nota y, al llegar a su límite, desplaza la matriz).
    def cargar_grid_proveedores(self, id_a_seleccionar=None):
        if not self.conn:
            return
        for widget in self.f_rows_dinamicas.winfo_children():
            widget.destroy()
        self.fila_matriz_seleccionada = None
        self.lista_widgets_filas = []
        c = self.conn.cursor()
        c.execute("SELECT id, categoria_suministro, proveedor_nombre, precio_lista, precio_descuento, precio_final_venta, notes_negociacion, notas_internas, cantidad FROM cotizacion_proveedores WHERE codigo_cotizacion = %s ORDER BY id ASC", (self.codigo_cot,))
        registros = c.fetchall()
        if not registros:
            self.actualizar_bloque_totales_pantalla()
            ctk.CTkLabel(self.f_rows_dinamicas, text="No hay costos asignados a este evento aún.", font=("Arial", 12, "italic"), text_color="#888").pack(pady=20)
            return
        self.actualizar_bloque_totales_pantalla()
        
        for i, r in enumerate(registros, start=1):
            id_real, cat_r, prov_r, pl_r, pd_r, pf_r, notas_r, notas_int_r, cant_r = r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], (r[8] if len(r) > 8 and r[8] else 1)
            cat_limpia = str(cat_r).strip("() '\",")
            prov_limpio = str(prov_r).strip("() '\",")
            data_pack = (id_real, cat_limpia, prov_limpio, pl_r, pd_r, pf_r, notas_r, cant_r)
            
            f_row = ctk.CTkFrame(self.f_rows_dinamicas, fg_color="#ffffff", border_width=1, border_color="#e0e0e0", corner_radius=0)
            f_row.pack(fill="x", pady=2)
            f_row.data_pack = data_pack
            self.lista_widgets_filas.append(f_row)
            
            row_idx = len(self.lista_widgets_filas) - 1

            def bind_drag(w):
                w.bind("<ButtonPress-1>", lambda e, idx=row_idx: self._iniciar_arrastre(e, idx), add="+")
                w.bind("<B1-Motion>", self._en_arrastre, add="+")
                w.bind("<ButtonRelease-1>", self._soltar_arrastre, add="+")

            bind_drag(f_row)

            def marcar_seleccion_f(event, f=f_row, d=data_pack):
                for child in self.f_rows_dinamicas.winfo_children():
                    child.configure(fg_color="#ffffff", border_color="#e0e0e0", border_width=1)
                    for sub in child.winfo_children():
                        if isinstance(sub, ctk.CTkTextbox):
                            if "FFFDE7" in sub.cget("fg_color"): continue 
                            sub.configure(fg_color="#ffffff")
                f.configure(fg_color="#cfe2ff")
                for sub in f.winfo_children():
                    if isinstance(sub, ctk.CTkTextbox):
                        if "FFFDE7" in sub.cget("fg_color"): continue 
                        sub.configure(fg_color="#cfe2ff")
                self.fila_matriz_seleccionada = d

            f_row.bind("<Button-1>", marcar_seleccion_f)
            
            # Ganancia por ítem = (P. Venta total) − (costo real unitario × cantidad).
            # El costo real es el P. Dscto si existe (> 0); si no, el P. Lista.
            costo_unit_real = float(pd_r) if pd_r and float(pd_r) > 0 else float(pl_r)
            ganancia_item = float(pf_r) - costo_unit_real * float(cant_r)

            anchos_row = [
                (str(i), 30, "center"), 
                (cat_limpia, 130, "w"), 
                (prov_limpio, 140, "w"), 
                (str(cant_r), 40, "center"), 
                (f"S/. {pl_r:.2f}", 70, "e"), 
                (f"S/. {pd_r:.2f}", 50, "e"), 
                (f"S/. {pf_r:.2f}", 80, "e"), 
                (f"S/. {ganancia_item:,.2f}", 85, "e", "#1e8449")
            ]
            
            for item in anchos_row:
                text, w, align = item[0], item[1], item[2]
                color = item[3] if len(item) > 3 else None
                just = "left" if align == "w" else ("right" if align == "e" else "center")
                wrap_val = w - 5 if w > 50 else 0 
                lbl = ctk.CTkLabel(f_row, text=text, font=("Arial", 11), width=w, wraplength=wrap_val, anchor=align, justify=just, text_color=color if color else None)
                lbl.pack(side="left", padx=2, fill="y")
                lbl.bind("<Button-1>", marcar_seleccion_f)
                bind_drag(lbl)
            
            # --- NOTAS DEL CLIENTE ---
            texto_nota = str(notas_r) if notas_r else "-"
            conteo_lineas = texto_nota.count('\n') + 1
            for linea_texto in texto_nota.split('\n'):
                conteo_lineas += len(texto_plano_sin_marcado(linea_texto)) // 65
                
            txt_notas = ctk.CTkTextbox(f_row, height=max(60, conteo_lineas * 20), font=("Helvetica", 10), fg_color="#ffffff", text_color="#000000", border_width=0, corner_radius=0, wrap="word")
            configurar_tags_formato(txt_notas, tam=10)
            insertar_texto_formateado(txt_notas, texto_nota)
            txt_notas.pack(side="left", fill="both", expand=True, padx=(4, 2), pady=5)
            
            txt_notas.bind("<Button-1>", lambda e, f=f_row, d=data_pack: marcar_seleccion_f(e, f, d))
            txt_notas._textbox.bind("<Button-1>", lambda e, f=f_row, d=data_pack: marcar_seleccion_f(e, f, d), add="+")
            bind_drag(txt_notas)
            bind_drag(txt_notas._textbox)
            
            # --- NOTAS INTERNAS ---
            texto_nota_int = str(notas_int_r) if notas_int_r else ""
            txt_notas_int = ctk.CTkTextbox(f_row, height=max(60, conteo_lineas * 20), font=("Helvetica", 10), fg_color="#FFFDE7", text_color="#000000", border_width=1, border_color="#FBC02D", corner_radius=5, wrap="word")
            txt_notas_int.insert("1.0", texto_nota_int)
            txt_notas_int.pack(side="left", fill="both", expand=True, padx=(2, 4), pady=5)
            
            txt_notas_int.bind("<Button-1>", lambda e, f=f_row, d=data_pack: marcar_seleccion_f(e, f, d))
            txt_notas_int._textbox.bind("<Button-1>", lambda e, f=f_row, d=data_pack: marcar_seleccion_f(e, f, d), add="+")
            bind_drag(txt_notas_int)
            bind_drag(txt_notas_int._textbox)

            if id_a_seleccionar == id_real:
                marcar_seleccion_f(None, f_row, data_pack)

    def mover_renglon_matriz(self, direccion):
        if not self.conn:
            return
        if not self.fila_matriz_seleccionada:
            return
        idx_act = -1
        for idx, widget in enumerate(self.lista_widgets_filas):
            if widget.data_pack == self.fila_matriz_seleccionada:
                idx_act = idx
                break
        if idx_act == -1:
            return
        idx_dest = idx_act - 1 if direccion == "ARRIBA" else idx_act + 1
        if idx_dest < 0 or idx_dest >= len(self.lista_widgets_filas):
            return
            
        id_act = self.lista_widgets_filas[idx_act].data_pack[0]
        id_dest = self.lista_widgets_filas[idx_dest].data_pack[0]
        
        c = self.conn.cursor()
        try:
            self._intercambiar_datos(id_act, id_dest, c)
            self.conn.commit()
            cache_sistema.invalidar()
        except Exception:
            self.conn.rollback()
        self.cargar_grid_proveedores(id_a_seleccionar=id_dest)


if __name__ == "__main__":
    pass